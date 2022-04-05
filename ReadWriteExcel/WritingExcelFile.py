from openpyxl import load_workbook

wb = load_workbook('WriteData.xlsx')
sheet = wb.active

# Write Column name
columns = sheet['A1'].value = "Name"
columns = sheet['B1'].value = "Age"
columns = sheet['C1'].value = "Salary"
wb.save('WriteData.xlsx')
print('Successfully write Column Name')

# insert data in columns
name = ['Muntasir', 'John', 'Smith']
sheet.cell(row=1, column=1).value = "Name"

j = 2
for i in range(0, 3):
    sheet.cell(row=j, column=1).value = name[i]
    j +=1

wb.save('WriteData.xlsx')
print('Successfully write Column Data')
